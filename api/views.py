from rest_framework.viewsets import ModelViewSet, GenericViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
import pandas as pd
import json
from .models import Teacher, Student, Room, ScheduleVersion, Group, RuleConfig
from .serializers import TeacherSerializer, StudentSerializer, RoomSerializer, ScheduleVersionSerializer, RuleConfigSerializer
from .utils import success_response, error_response

# ==================== RuleConfig 视图 ====================
class RuleConfigViewSet(ModelViewSet):
    queryset = RuleConfig.objects.all()
    serializer_class = RuleConfigSerializer

# ==================== 导入混入类 ====================
class ImportMixin:
    @action(detail=False, methods=['post'])
    def import_data(self, request):
        file = request.FILES.get('file')
        if not file:
            return error_response(message='未提供文件', status=400)

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file)
            else:
                return error_response(message='不支持的文件格式', status=400)

            df.columns = [c.strip() for c in df.columns]
            df = df.where(pd.notnull(df), None)
            data_list = df.to_dict(orient='records')

            success_count = 0
            errors = []

            for index, row in enumerate(data_list):
                try:
                    processed_row = {}
                    for k, v in row.items():
                        processed_row[k] = v
                        snake_k = ''.join(['_' + i.lower() if i.isupper() else i for i in k]).lstrip('_')
                        if snake_k not in processed_row:
                            processed_row[snake_k] = v

                    json_fields = ['roles', 'available_types', 'availableTypes', 'defense_types', 'defenseTypes']
                    for field in json_fields:
                        if field in processed_row and isinstance(processed_row[field], str):
                            val = processed_row[field].strip()
                            if not val:
                                processed_row[field] = []
                            elif (val.startswith('[') and val.endswith(']')) or (val.startswith('{') and val.endswith('}')):
                                try:
                                    processed_row[field] = json.loads(val.replace("'", '"'))
                                except:
                                    processed_row[field] = [item.strip() for item in val.split(',') if item.strip()]
                            else:
                                processed_row[field] = [item.strip() for item in val.split(',') if item.strip()]

                    bool_fields = ['isExternal', 'is_external']
                    for field in bool_fields:
                        if field in processed_row and isinstance(processed_row[field], str):
                            processed_row[field] = processed_row[field].lower() in ['true', '1', '是', 'yes']

                    serializer = self.get_serializer(data=processed_row)
                    if serializer.is_valid():
                        serializer.save()
                        success_count += 1
                    else:
                        errors.append(f"行 {index + 2}: {serializer.errors}")
                except Exception as e:
                    errors.append(f"行 {index + 2}: {str(e)}")

            if success_count == 0 and errors:
                return error_response(message='导入失败，请检查文件格式', status=400, errors=errors[:5])

            return success_response(
                data={'success_count': success_count, 'errors': errors},
                message=f'成功导入 {success_count} 条数据',
                status=200
            )

        except Exception as e:
            return error_response(message=f'解析文件失败: {str(e)}', status=400)

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return error_response(message='未提供待删除的 ID 列表', status=400)

        try:
            queryset = self.get_queryset().filter(id__in=ids)
            count = queryset.count()
            queryset.delete()
            return success_response(data={'deleted_count': count}, message=f'成功删除 {count} 条数据', status=200)
        except Exception as e:
            return error_response(message=f'删除失败: {str(e)}', status=400)

# ==================== 教师视图 ====================
class TeacherViewSet(ImportMixin, ModelViewSet):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return success_response(data=serializer.data, message='获取教师列表成功', status=200)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='添加教师成功', status=201)
        return error_response(message='添加教师失败', status=400, errors=serializer.errors)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='更新教师成功', status=200)
        return error_response(message='更新教师失败', status=400, errors=serializer.errors)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='删除教师成功', status=204)

# ==================== 学生视图 ====================
class StudentViewSet(ImportMixin, ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return success_response(data=serializer.data, message='获取学生列表成功', status=200)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='添加学生成功', status=201)
        return error_response(message='添加学生失败', status=400, errors=serializer.errors)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='更新学生成功', status=200)
        return error_response(message='更新学生失败', status=400, errors=serializer.errors)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='删除学生成功', status=204)

# ==================== 教室视图 ====================
class RoomViewSet(ImportMixin, ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return success_response(data=serializer.data, message='获取教室列表成功', status=200)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='添加教室成功', status=201)
        return error_response(message='添加教室失败', status=400, errors=serializer.errors)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return success_response(data=serializer.data, message='更新教室成功', status=200)
        return error_response(message='更新教室失败', status=400, errors=serializer.errors)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return success_response(message='删除教室成功', status=204)

# ==================== 排期视图 ====================
class ScheduleViewSet(GenericViewSet):
    queryset = ScheduleVersion.objects.all()
    serializer_class = ScheduleVersionSerializer

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """一键生成排期"""
        teachers = Teacher.objects.all().values()
        students = Student.objects.all().values()
        rooms = Room.objects.all().values()

        rules = request.data.get('rules', {
            'defense_type': 'pre',
            'start_date': '2025-05-10',
            'avoid_weekend': True,
            'avoid_holiday': True,
        })

        input_data = {
            'teachers': list(teachers),
            'students': list(students),
            'rooms': list(rooms),
            'rules': rules
        }

        try:
            from algorithm import generate_schedule
            result = generate_schedule(**input_data)
        except ImportError:
            result = self._mock_schedule_result(input_data)

        ScheduleVersion.objects.filter(defense_type=rules['defense_type']).update(is_current=False)

        version_num = ScheduleVersion.objects.filter(defense_type=rules['defense_type']).count() + 1
        schedule_version = ScheduleVersion.objects.create(
            version=version_num,
            defense_type=rules['defense_type'],
            rules_snapshot=rules,
            is_current=True
        )

        for group_data in result.get('groups', []):
            group = Group.objects.create(
                schedule_version=schedule_version,
                group_id=group_data['group_id'],
                time=group_data['time'],
                room_id=group_data.get('room_id'),
                campus=group_data.get('campus', ''),
                chair_id=group_data.get('chair_id'),
                secretary_id=group_data.get('secretary_id')
            )
            for expert_id in group_data.get('expert_ids', []):
                group.experts.through.objects.create(group_id=group.id, teacher_id=expert_id)
            for student_id in group_data.get('student_ids', []):
                group.students.through.objects.create(group_id=group.id, student_id=student_id)

        request.query_params._mutable = True
        request.query_params['defense_type'] = rules['defense_type']
        current_response = self.current(request)
        if hasattr(current_response, 'data'):
            data = current_response.data
        else:
            data = current_response

        return success_response(data=data, message='排期生成成功', status=201)

    def _mock_schedule_result(self, input_data):
        return {
            'groups': [
                {
                    'group_id': 'G1',
                    'time': '2025-05-10 09:00-12:00',
                    'room_id': 1,
                    'campus': '创新港',
                    'chair_id': 1,
                    'expert_ids': [2, 3],
                    'secretary_id': 4,
                    'student_ids': [101, 102, 103]
                }
            ],
            'conflicts': []
        }

    @action(detail=False, methods=['get'])
    def current(self, request):
        """获取当前生效的排期"""
        defense_type = request.query_params.get('defense_type', 'pre')
        schedule_version = ScheduleVersion.objects.filter(
            defense_type=defense_type,
            is_current=True
        ).first()

        if not schedule_version:
            return success_response(
                data={'defenseType': defense_type, 'generatedAt': '', 'groups': []},
                message='暂无排期结果',
                status=200
            )

        groups = schedule_version.groups.all()
        formatted_groups = []
        for group in groups:
            time_parts = group.time.split(' ')
            date = time_parts[0] if len(time_parts) > 0 else ''
            time_range = time_parts[1] if len(time_parts) > 1 else ''

            formatted_groups.append({
                'id': group.id,
                'defenseType': schedule_version.get_defense_type_display(),
                'groupName': group.group_id,
                'campus': group.campus,
                'classroom': group.room.name if group.room else '未分配',
                'date': date,
                'timeRange': time_range,
                'chairman': group.chair.name if group.chair else None,
                'secretary': group.secretary.name if group.secretary else '未分配',
                'teachers': [
                    {'id': t.id, 'name': t.name, 'title': t.title, 'roles': getattr(t, 'roles', [])}
                    for t in group.experts.all()
                ],
                'students': [
                    {
                        'id': s.id,
                        'name': s.name,
                        'studentType': s.student_type,
                        'mentorName': s.mentor_name
                    }
                    for s in group.students.all()
                ]
            })

        data = {
            'defenseType': schedule_version.get_defense_type_display(),
            'generatedAt': schedule_version.created_at.strftime('%Y-%m-%d %H:%M'),
            'groups': formatted_groups
        }
        return success_response(data=data, message='获取当前排期成功', status=200)

    @action(detail=False, methods=['post'])
    def adjust(self, request):
        """人工调整：移动学生、更换专家、修改时间/教室"""
        action_type = request.data.get('action')

        if action_type == 'move_student':
            return self._move_student(request)
        elif action_type == 'change_expert':
            return self._change_expert(request)
        elif action_type == 'change_time':
            return self._change_time(request)
        elif action_type == 'change_room':
            return self._change_room(request)
        elif action_type == 'change_chair':
            return self._change_chair(request)
        elif action_type == 'change_secretary':
            return self._change_secretary(request)
        else:
            return error_response(message=f'未知的 action: {action_type}', status=400)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出当前排期为 Excel 文件（直接返回文件，不包装）"""
        defense_type = request.query_params.get('defense_type', 'pre')
        schedule_version = ScheduleVersion.objects.filter(
            defense_type=defense_type,
            is_current=True
        ).first()

        if not schedule_version:
            return error_response(message='暂无排期结果可导出', status=404)

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        groups = schedule_version.groups.all()

        colors = [
            'FFB3B3', 'B3FFB3', 'B3B3FF', 'FFFFB3', 'FFB3FF', 'B3FFFF',
        ]

        mentor_color_map = {}
        color_index = 0

        for group in groups:
            sheet = wb.create_sheet(title=f"组{group.group_id}")

            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 25

            title_font = Font(bold=True, size=14)
            title_cell = sheet['A1']
            title_cell.value = f"答辩排期表 - {group.group_id}"
            title_cell.font = title_font
            sheet.merge_cells('A1:D1')

            row = 3
            info_data = [
                ('时间', group.time),
                ('教室', group.room.name if group.room else '未分配'),
                ('校区', group.campus),
                ('主席/组长', group.chair.name if group.chair else '未分配'),
                ('秘书', group.secretary.name if group.secretary else '未分配'),
            ]

            for label, value in info_data:
                sheet[f'A{row}'] = label
                sheet[f'B{row}'] = value
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1

            sheet[f'A{row}'] = '专家'
            sheet[f'A{row}'].font = Font(bold=True)
            expert_names = [e.name for e in group.experts.all()]
            sheet[f'B{row}'] = '、'.join(expert_names) if expert_names else '未分配'
            row += 2

            headers = ['学生姓名', '学生类型', '导师姓名', '导师职称']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

            row += 1

            for student in group.students.all():
                mentor_key = student.mentor_name if student.mentor_name else '无导师'
                if mentor_key not in mentor_color_map:
                    mentor_color_map[mentor_key] = colors[color_index % len(colors)]
                    color_index += 1

                color = mentor_color_map[mentor_key]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                cell = sheet.cell(row=row, column=1, value=student.name)
                cell.fill = fill

                cell = sheet.cell(row=row, column=2, value=student.student_type)
                cell.fill = fill

                cell = sheet.cell(row=row, column=3, value=student.mentor_name if student.mentor_name else '未分配')
                cell.fill = fill

                cell = sheet.cell(row=row, column=4, value='')
                cell.fill = fill

                row += 1

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for r in range(3, row):
                for c in range(1, 5):
                    cell = sheet.cell(row=r, column=c)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=defense_schedule_{defense_type}.xlsx'
        wb.save(response)
        return response

    # ========== 独立冲突检测接口 ==========
    @action(detail=False, methods=['post'])
    def check_conflicts(self, request):
        """
        独立冲突检测接口
        请求体: {"schedule_version_id": 123}
        返回该版本下的所有冲突
        """
        version_id = request.data.get('schedule_version_id')
        if not version_id:
            return error_response(message='缺少 schedule_version_id', status=400)
        try:
            schedule_version = ScheduleVersion.objects.get(id=version_id)
            conflicts = self._check_conflicts(schedule_version)
            return success_response(data={'conflicts': conflicts}, message='检测成功')
        except ScheduleVersion.DoesNotExist:
            return error_response(message='排期版本不存在', status=404)

    # ---------- 私有调整方法 ----------
    def _move_student(self, request):
        student_id = request.data.get('student_id')
        from_group_id = request.data.get('from_group_id')
        to_group_id = request.data.get('to_group_id')

        if not all([student_id, from_group_id, to_group_id]):
            return error_response(message='缺少必要参数', status=400)

        try:
            from_group = Group.objects.get(id=from_group_id)
            to_group = Group.objects.get(id=to_group_id)

            from_group.students.remove(student_id)
            to_group.students.add(student_id)

            conflicts = self._check_conflicts(from_group.schedule_version)
            return success_response(
                data={'conflicts': conflicts},
                message=f'学生 {student_id} 已从组 {from_group_id} 移动到组 {to_group_id}',
                status=200
            )
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)
        except Exception as e:
            return error_response(message=str(e), status=400)

    def _change_expert(self, request):
        group_id = request.data.get('group_id')
        old_expert_id = request.data.get('old_expert_id')
        new_expert_id = request.data.get('new_expert_id')

        if not all([group_id, old_expert_id, new_expert_id]):
            return error_response(message='缺少必要参数', status=400)

        try:
            group = Group.objects.get(id=group_id)
            group.experts.remove(old_expert_id)
            group.experts.add(new_expert_id)
            conflicts = self._check_conflicts(group.schedule_version)
            return success_response(data={'conflicts': conflicts}, message='专家已更换', status=200)
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)

    def _change_chair(self, request):
        group_id = request.data.get('group_id')
        new_chair_id = request.data.get('new_chair_id')

        if not all([group_id, new_chair_id]):
            return error_response(message='缺少必要参数', status=400)

        try:
            group = Group.objects.get(id=group_id)
            group.chair_id = new_chair_id
            group.save()
            conflicts = self._check_conflicts(group.schedule_version)
            return success_response(data={'conflicts': conflicts}, message='主席已更换', status=200)
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)

    def _change_secretary(self, request):
        group_id = request.data.get('group_id')
        new_secretary_id = request.data.get('new_secretary_id')

        if not all([group_id, new_secretary_id]):
            return error_response(message='缺少必要参数', status=400)

        try:
            group = Group.objects.get(id=group_id)
            group.secretary_id = new_secretary_id
            group.save()
            conflicts = self._check_conflicts(group.schedule_version)
            return success_response(data={'conflicts': conflicts}, message='秘书已更换', status=200)
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)

    def _change_time(self, request):
        group_id = request.data.get('group_id')
        new_time = request.data.get('new_time')

        if not all([group_id, new_time]):
            return error_response(message='缺少必要参数', status=400)

        try:
            group = Group.objects.get(id=group_id)
            group.time = new_time
            group.save()
            conflicts = self._check_conflicts(group.schedule_version)
            return success_response(data={'conflicts': conflicts}, message=f'时间已修改为 {new_time}', status=200)
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)

    def _change_room(self, request):
        group_id = request.data.get('group_id')
        new_room_id = request.data.get('new_room_id')

        if not all([group_id, new_room_id]):
            return error_response(message='缺少必要参数', status=400)

        try:
            group = Group.objects.get(id=group_id)
            group.room_id = new_room_id
            group.save()
            conflicts = self._check_conflicts(group.schedule_version)
            return success_response(data={'conflicts': conflicts}, message='教室已更换', status=200)
        except Group.DoesNotExist:
            return error_response(message='组不存在', status=404)

    def _check_conflicts(self, schedule_version):
        conflicts = []
        groups = schedule_version.groups.all()

        teacher_time_map = {}

        for group in groups:
            if group.chair:
                key = (group.chair.id, group.time)
                if key in teacher_time_map:
                    conflicts.append({
                        'type': 'teacher_time_conflict',
                        'description': f"{group.chair.name} 在同一时间 {group.time} 被分配到多个组",
                        'teacher_id': group.chair.id,
                        'teacher_name': group.chair.name,
                        'time': group.time,
                        'group_ids': [teacher_time_map[key], group.id]
                    })
                else:
                    teacher_time_map[key] = group.id

            for expert in group.experts.all():
                key = (expert.id, group.time)
                if key in teacher_time_map:
                    conflicts.append({
                        'type': 'teacher_time_conflict',
                        'description': f"{expert.name} 在同一时间 {group.time} 被分配到多个组",
                        'teacher_id': expert.id,
                        'teacher_name': expert.name,
                        'time': group.time,
                        'group_ids': [teacher_time_map[key], group.id]
                    })
                else:
                    teacher_time_map[key] = group.id

            if group.secretary:
                key = (group.secretary.id, group.time)
                if key in teacher_time_map:
                    conflicts.append({
                        'type': 'teacher_time_conflict',
                        'description': f"{group.secretary.name} 在同一时间 {group.time} 被分配到多个组",
                        'teacher_id': group.secretary.id,
                        'teacher_name': group.secretary.name,
                        'time': group.time,
                        'group_ids': [teacher_time_map[key], group.id]
                    })
                else:
                    teacher_time_map[key] = group.id

        room_time_map = {}
        for group in groups:
            if group.room:
                key = (group.room.id, group.time)
                if key in room_time_map:
                    conflicts.append({
                        'type': 'room_conflict',
                        'description': f"{group.room.name} 在同一时间 {group.time} 被多个组使用",
                        'room_id': group.room.id,
                        'room_name': group.room.name,
                        'time': group.time,
                        'group_ids': [room_time_map[key], group.id]
                    })
                else:
                    room_time_map[key] = group.id

        # 秘书-学生冲突
        for group in groups:
            if group.secretary:
                secretary_name = group.secretary.name
                for student in group.students.all():
                    if student.secretary_name == secretary_name:
                        conflicts.append({
                            'type': 'secretary_student_conflict',
                            'description': f"秘书 {secretary_name} 和自己的学生 {student.name} 在同一组",
                            'group_id': group.id,
                            'secretary_name': secretary_name,
                            'student_id': student.id
                        })

        return conflicts